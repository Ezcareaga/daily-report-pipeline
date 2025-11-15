#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel file generation utilities."""

from pathlib import Path
from typing import List, Any, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from .config import ConfigManager
from .exceptions import PipelineError


class ExcelGenerator:
    """Handles Excel file generation with custom formatting."""
    
    def __init__(self, config: Optional[ConfigManager] = None):
        """
        Initialize Excel generator.
        
        Args:
            config: Optional configuration manager
        """
        self.config = config
        
        if config:
            self.number_format = config.get('ARCHIVOS', 'formato_numero', default='europeo')
            self.decimals = config.getint('ARCHIVOS', 'decimales', default=2)
        else:
            self.number_format = 'europeo'
            self.decimals = 2
    
    def get_number_format_string(self) -> str:
        """
        Get Excel number format string based on configuration.
        
        Returns:
            str: Excel number format string
        """
        if self.number_format == 'europeo':
            if self.decimals == 0:
                return '#.##0'
            elif self.decimals == 1:
                return '#.##0,0'
            elif self.decimals == 2:
                return '#.##0,00'
            else:
                return '#.##0,' + '0' * self.decimals
        else:
            if self.decimals == 0:
                return '#,##0'
            elif self.decimals == 1:
                return '#,##0.0'
            elif self.decimals == 2:
                return '#,##0.00'
            else:
                return '#,##0.' + '0' * self.decimals

    def create_workbook(
        self,
        data: List[List[Any]],
        headers: Optional[List[str]] = None,
        sheet_name: str = "Reporte"
    ) -> Workbook:
        """
        Create Excel workbook from data.
        
        Args:
            data: List of rows (each row is a list of values)
            headers: Optional column headers
            sheet_name: Name for the worksheet
            
        Returns:
            Workbook: openpyxl Workbook object
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        current_row = 1
        num_format = self.get_number_format_string()
        
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
            current_row += 1
        
        for row_data in data:
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = num_format
                elif isinstance(value, datetime):
                    cell.value = value
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                else:
                    cell.value = str(value) if value is not None else ''
            
            current_row += 1
        
        return wb
    
    def save_workbook(self, workbook: Workbook, file_path: Path) -> None:
        """
        Save workbook to file.
        
        Args:
            workbook: openpyxl Workbook
            file_path: Path where to save
            
        Raises:
            PipelineError: If save fails
        """
        try:
            file_path = Path(file_path)
            file_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(file_path)
        except Exception as e:
            raise PipelineError(f"Failed to save Excel file: {e}") from e